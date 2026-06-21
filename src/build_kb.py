"""
LangChain Knowledge Base Builder for RFP Intelligence.

This script builds the RAG knowledge base from ZIP-based source folders.
It extracts content from company PDFs, the USask DOCX proposal, and selected
Excel tender scopes, converts the extracted content into LangChain Document
objects, saves kb_documents.json, and optionally rebuilds the Chroma index.

Processing flow:
    source ZIP files
    -> file-specific parsing
    -> cleaning / optional masking / capability extraction
    -> chunking with SmartChunker
    -> LangChain Document(page_content, metadata)
    -> JSON export
    -> Chroma vector index

This version includes timing logs for each pipeline and the full build.

Run:
    python src/build_kb_langchain.py --reset
    python src/build_kb_langchain.py --no-index
"""

from __future__ import annotations
import openpyxl
import argparse
import hashlib
import json
import os
import re
import tempfile
import zipfile
import time
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
import multiprocessing
from tqdm import tqdm  # pip install tqdm

import pdfplumber
from dotenv import load_dotenv
from langchain_core.documents import Document
from openai import OpenAI
from src.vectorstores import ChromaVectorStore

try:
    from docling.document_converter import DocumentConverter
except Exception:
    DocumentConverter = None

from src import config
from src.chunking_langchain import SmartChunker

load_dotenv()

# =============================================================================
# 1) PATHS / SETTINGS
# =============================================================================

ROOT_DIR = Path(__file__).resolve().parent.parent
SOURCE_DATA_DIR = ROOT_DIR / "_source_data"
COMPANY_ZIP = SOURCE_DATA_DIR / "Beam_WeCloud_RAG_Documents.zip"
TENDERS_ZIP = SOURCE_DATA_DIR / "Tenders - Raw and Annotated.zip"
DATA_DIR = ROOT_DIR / "data_langchain"
DOCS_JSON = DATA_DIR / "kb_documents.json"
CHROMA_DIR = DATA_DIR / "chroma_db"
USASK_RFP_ID = "USask-GenAI-730126"

SCRIPT_VERSION = "v2.1.0-parallel-processing-timing"
PROCESSING_TIMESTAMP = datetime.now(timezone.utc).isoformat()

CHUNKER = SmartChunker(
    chunk_size=config.CHUNK_SIZE,
    chunk_overlap=config.CHUNK_OVERLAP,
    min_length=config.MIN_CHUNK_LENGTH,
)

# =============================================================================
# 1.1) PARALLEL PROCESSING CONFIGURATION
# =============================================================================

"""
Worker configuration used by the parallel pipelines.

- ProcessPoolExecutor is used for PDF parsing tasks that run as separate
  processes. Each submitted task receives the PDF bytes and returns a list of
  LangChain Documents.
- ThreadPoolExecutor is used for Excel scope files because each Excel task calls
  the LLM to generate a capability summary. Each submitted task returns the
  extracted capability payload, which is converted into Documents afterward.

The values below are printed at startup so the build log shows the runtime
configuration used for the experiment.
"""

# Runtime worker settings
CPU_CORES = multiprocessing.cpu_count()
MAX_WORKERS_CPU = CPU_CORES  # processes for PDF workers
MAX_WORKERS_IO = min(20, CPU_CORES * 2)  # threads for Excel / LLM workers
DOCX_SECTION_WORKERS = min(8, MAX_WORKERS_IO)  # threads for DOCX section masking/tagging

print(f"🔧 Parallel Processing Configuration:")
print(f"   - CPU cores available: {CPU_CORES}")
print(f"   - ProcessPoolExecutor workers: {MAX_WORKERS_CPU} (CPU-bound)")
print(f"   - ThreadPoolExecutor workers: {MAX_WORKERS_IO} (I/O-bound)")
print(f"   - DOCX section workers: {DOCX_SECTION_WORKERS} (LLM masking/tagging)")

# =============================================================================
# 2) FILE RULES
# =============================================================================

"""
File rules used by the ingestion pipelines.
Each entry maps a source file to the metadata and tags that should be attached
when its chunks are converted into LangChain Documents.
"""

COMPANY_PDF_FILES: Dict[str, Dict[str, object]] = {
    "Beamdata Past Project Descriptions.pdf": {
        "doc_type": "past_projects",
        "document_tags": ["past-project", "weclouddata", "beamdata", "company"],
        "pipeline_name": "company_pdf_past_projects_pipeline",
        "prefix": "Note: WeCloudData is a sister company of Beamdata. Projects appear under both brands.\n\n",
    },
    "Beam Data AI Hub Intro.pptx (1).pdf": {
        "doc_type": "company_intro",
        "document_tags": ["beamdata", "company", "genai", "product", "ai-hub"],
        "pipeline_name": "company_pdf_ai_hub_intro_pipeline",
        "prefix": "",
    },
}

PDF_PAGE_OVERRIDES: Dict[str, Dict[str, List[int]]] = {
    "Beam Data AI Hub Intro.pptx (1).pdf": {
        "skip_pages": [1, 2, 6, 7, 13, 24, 25, 26, 27, 28],
    },
}

DOCX_SECTION_OVERRIDES: Dict[str, Dict[str, object]] = {
    "Copy of CP-730126 - Generative Artificial Intelligence (AI) Software.docx": {
        "stop_at_section": "Tender Summary",
        "skip_sections": [
            "table of contents",
            "submission checklist",
            "questions to tender",
            "appendix",
            "pricing",
            "software as a service vendor form",
            "section planning guide",
            "point system",
            "point weighting",
            "key dates",
            "document to summit",
            "things need to mention",
            "reseach",
            "research",
            "remark",
        ],
    },
}

DEFAULT_DOCX_SKIP_SECTION_PATTERNS = {
    "table of contents",
    "tender summary",
    "submission checklist",
    "questions to tender",
    "appendix",
    "pricing",
    "software as a service vendor form",
    "section planning guide",
    "point system",
    "point weighting",
    "key dates",
    "document to summit",
    "things need to mention",
    "reseach",
    "research",
    "remark",
}

# Target Excel tender files for capability extraction
TENDER_EXCEL_TARGET_FILES = {
    "usask - genai software",
    "ttc - implementation of a high availability sql database solution",
    "scaleai - digital intelligence training grants for canadian businesses",
    "nrfp 5549 - learning design and authoring tool",
    "icbc bid planner - learning design and authoring tool",
    "glenbow-alberta institute - website development",
    "design and implementation of an ai-powered data platform for pakistan customs",
}

# =============================================================================
# 3) HELPER FUNCTIONS
# =============================================================================

def normalize_file_stem(name: str) -> str:
    """Return a normalized filename stem for matching Excel targets."""
    stem = Path(name).stem.lower().strip()
    stem = re.sub(r"\s+", " ", stem)
    return stem

def slugify_tag(value: Any) -> str:
    """Convert a capability or industry value into a safe metadata tag."""
    tag = str(value or "").strip().lower()
    tag = tag.replace("/", "-").replace("&", "and")
    tag = re.sub(r"[^a-z0-9]+", "-", tag)
    tag = re.sub(r"-+", "-", tag).strip("-")
    return tag

# =============================================================================
# 4) TAGGING SYSTEM
# =============================================================================

"""
Tagging helpers.
Auto-tagging scans chunk text for known keywords. LLM tagging optionally adds
context-aware tags from the same controlled tag list.
"""

_TAG_MAP: Dict[str, List[str]] = {
    "genai": ["generative ai", "large language model", "llm", "gpt", "chatgpt", "genai", "foundation model"],
    "ml": ["machine learning", "deep learning", "neural network", "model training", "ai model"],
    "rag": ["retrieval augmented", "vector search", "embedding", "semantic search", "knowledge base"],
    "cloud": ["aws", "amazon web services", "azure", "google cloud", "gcp", "cloud infrastructure"],
    "microsoft": ["microsoft", "microsoft 365", "m365", "sharepoint", "teams", "azure ad", "copilot"],
    "data-pipeline": ["data pipeline", "etl", "data ingestion", "apache spark", "airflow", "data engineering"],
    "nlp": ["natural language processing", "nlp", "text analysis", "sentiment", "text classification"],
    "automation": ["automation", "automated", "rpa", "workflow automation", "process automation"],
    "government": ["government", "public sector", "municipal", "federal", "crown corporation", "procurement"],
    "education": ["university", "college", "student", "academic", "learning", "curriculum", "campus"],
    "utility": ["utility", "bcuc", "bc utilities", "energy", "hydro", "grid", "regulation"],
    "healthcare": ["health", "medical", "patient", "clinical", "hospital", "pharmacy"],
    "finance": ["financial", "banking", "payment", "billing", "invoice", "accounting"],
    "past-project": ["project description", "past project", "case study", "client engagement", "delivered"],
    "proposal": ["proposal", "we propose", "our approach", "beamdata will", "our team will"],
    "lms": ["learning management", "lms", "course", "training platform", "e-learning"],
    "compliance": ["compliance", "governance", "regulation", "audit", "policy", "gdpr", "privacy"],
    "migration": ["migration", "migrate", "transition", "modernization", "upgrade"],
    "implementation": ["implementation", "deploy", "rollout", "integration", "onboarding"],
    "data-governance": ["data governance", "data quality", "master data", "data catalog", "lineage"],
}

def auto_tag(text: str, base_tags: Optional[List[str]] = None) -> List[str]:
    """Add keyword-based tags to a text block using the controlled tag map."""
    tags = set(base_tags or [])
    t = text.lower()
    for tag, keywords in _TAG_MAP.items():
        if any(keyword in t for keyword in keywords):
            tags.add(tag)
    return sorted(tags)

def llm_tag(text: str, client: OpenAI, base_tags: Optional[List[str]] = None) -> List[str]:
    """Ask the LLM to add up to five controlled tags for the given text."""
    base = auto_tag(text, base_tags)
    tag_list = ", ".join(sorted(_TAG_MAP.keys()))
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.1,  # low temperature for stable tag output
            max_tokens=80,
            messages=[{
                "role": "system",
                "content": "You are an expert tagger for technical documents. Choose the most relevant tags."
            }, {
                "role": "user",
                "content": (
                    f"Choose up to 5 relevant tags from this list for the text below.\n"
                    f"Tags: {tag_list}\n"
                    f"Return ONLY a comma-separated list of tag names.\n\n"
                    f"Text: {text[:600]}"
                ),
            }],
        )
        raw = resp.choices[0].message.content.strip().lower()
        llm_tags = [t.strip() for t in raw.split(",") if t.strip() in _TAG_MAP]
        return sorted(set(base) | set(llm_tags))
    except Exception:
        return base

# =============================================================================
# 5) SENSITIVE DATA OBFUSCATION
# =============================================================================

"""
Sensitive-data obfuscation helpers.
The LLM masking step removes names, client names, addresses, and similar
sensitive values. The regex pass then catches common structured values such as
emails, phone numbers, and URLs.
"""

_MASKING_WHITELIST_STR = (
    "Beamdata, WeCloudData, Beam Data, We Cloud Data, BeamData AI, "
    "Sportradar, WCC, Globe and Mail, TrustABC, XYZ Robotics, "
    "Saudi Digital Academy, MCIT, HRDF, AI71, Ministry of Defense, "
    "Singapore Management University, SIT, StackFuel"
)

def _regex_validate_and_fix(text: str) -> str:
    """Replace emails, phone numbers, and URLs with standard placeholders."""
    patterns = [
        (r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", "[EMAIL]"),
        (r"(?:\+?\d[\d\-\s()]{7,}\d)", "[PHONE]"),
        (r"https?://\S+|www\.\S+", "[URL]"),
    ]
    for pattern, replacement in patterns:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    return text

def _mask_via_llm(text: str, client: OpenAI) -> str:
    """Use the LLM to obfuscate sensitive values while preserving the text meaning."""
    prompt = (
        "You are a sensitive-data obfuscation tool.\n\n"
        "Obfuscate any sensitive information in the text below using these placeholders:\n"
        "- Personal names                      -> [PERSON_NAME]\n"
        "- Client or company names              -> [CLIENT_NAME]\n"
        "- Email addresses                      -> [EMAIL]\n"
        "- Phone numbers                        -> [PHONE]\n"
        "- Physical addresses                   -> [ADDRESS]\n"
        "- URLs                                  -> [URL]\n"
        "- Identification numbers or IDs         -> [ID]\n"
        "- Postal codes                          -> [POSTAL_CODE]\n\n"
        f"Whitelist: never obfuscate these names: {_MASKING_WHITELIST_STR}\n\n"
        "Rules:\n"
        "- Return only the obfuscated text. Do not add explanations.\n"
        "- Do not change technical or business context unless it is sensitive.\n"
        "- Preserve the original structure and formatting as much as possible.\n\n"
        f"Text:\n{text}"
    )
    try:
        resp = client.chat.completions.create(
            model=config.LLM_MODEL,
            temperature=0,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        print(f"  ⚠️  LLM masking failed: {e} — keeping original text")
        return text

# =============================================================================
# 6) TEXT CLEANING HELPERS
# =============================================================================

FOOTER_PATTERNS = [
    r"(?m)^Page \d+\s+of\s+\d+\s*$",
    r"(?m)^\s*\d{1,3}\s*$",
    r"(?m)^Confidential\s*$",
    r"(?m)^CONFIDENTIAL\s*$",
]

def clean(text: str) -> str:
    """Remove repeated footer/header noise and normalize whitespace."""
    for pattern in FOOTER_PATTERNS:
        text = re.sub(pattern, "", text)
    text = re.sub(r"[-_]{8,}", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()

def sha256_bytes(data: bytes) -> str:
    """Return SHA-256 for raw bytes."""
    return hashlib.sha256(data).hexdigest()

def file_hash(path: Path) -> str:
    """Return SHA-256 for a file on disk."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()

def resolve_zip_path(preferred_path: Path, base_name: str) -> Optional[Path]:
    """Find a source ZIP by the configured path or by matching the base name."""
    if preferred_path.exists():
        return preferred_path
    candidates = list(SOURCE_DATA_DIR.glob(f"{base_name}*.zip"))
    candidates += [p for p in SOURCE_DATA_DIR.glob(f"{base_name}*") if p.is_file() and p.suffix.lower() == ".zip"]
    return candidates[0] if candidates else None

def find_zip_entry(zf: zipfile.ZipFile, target_file_name: str) -> Optional[str]:
    """Find a target file inside an opened ZIP archive by filename."""
    target_lower = target_file_name.lower()
    for entry in zf.infolist():
        if entry.is_dir():
            continue
        if Path(entry.filename).name.lower() == target_lower:
            return entry.filename
    return None

def safe_metadata(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """Convert metadata values into Chroma-safe scalar strings/numbers/bools."""
    safe: Dict[str, Any] = {}
    for key, value in metadata.items():
        if value is None:
            safe[key] = ""
        elif isinstance(value, (str, int, float, bool)):
            safe[key] = value
        else:
            safe[key] = json.dumps(value, ensure_ascii=False)
    return safe

def build_document_metadata(
    *,
    source: str,
    source_folder: str,
    doc_type: str,
    pipeline_name: str,
    document_tags: List[str],
    file_hash_value: Optional[str],
    rfp_id: Optional[str] = None,
    is_proposal: bool = False,
    content_type: Optional[str] = None,
    file_ext: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Build standard document metadata structure.
    
    This ensures consistent metadata across all document types.
    All documents include processing timestamp and script version.
    
    Args:
        source: Original file name
        source_folder: ZIP file name
        doc_type: Type of document
        pipeline_name: Name of processing pipeline
        document_tags: List of tags
        file_hash_value: SHA-256 hash of file
        rfp_id: RFP identifier (if applicable)
        is_proposal: Whether this is a proposal document
        content_type: Type of content
        file_ext: File extension
        
    Returns:
        Metadata dictionary
    """
    return {
        "source": source,
        "source_folder": source_folder,
        "doc_type": doc_type,
        "content_type": content_type or "",
        "pipeline_name": pipeline_name,
        "file_ext": file_ext or "",
        "file_hash": file_hash_value or "",
        "rfp_id": rfp_id or "",
        "is_proposal": is_proposal,
        "document_tags": sorted(set(document_tags)),
        "processed_at": PROCESSING_TIMESTAMP,
        "script_version": SCRIPT_VERSION,
    }

def make_chunk_id(
    document_metadata: Dict[str, Any],
    chunk_index: int,
    page: Optional[int] = None,
    section: str = ""
) -> str:
    """
    Generate a unique chunk ID.
    
    The ID is based on:
    - Document source
    - File hash (ensures uniqueness across runs)
    - Page/section information
    - Chunk index
    
    Args:
        document_metadata: Document metadata
        chunk_index: Index of this chunk
        page: Page number (for PDFs)
        section: Section name (for DOCX)
        
    Returns:
        SHA-1 hash as string
    """
    raw = "|".join([
        str(document_metadata.get("source", "")),
        str(document_metadata.get("file_hash", "")),
        str(page or ""),
        section or "",
        str(chunk_index),
    ])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

# =============================================================================
# 7) DOCUMENT BUILDER
# =============================================================================

class KBDocumentBuilder:
    """Convert cleaned chunk text into LangChain Document objects."""
    
    def build(
        self,
        *,
        content: str,
        document_metadata: Dict[str, Any],
        chunk_index: int,
        chunk_tags: List[str],
        page: Optional[int] = None,
        section: str = "",
    ) -> Document:
        """
        Build a single LangChain Document.
        
        Args:
            content: Text content of the chunk
            document_metadata: Base metadata from the document
            chunk_index: Index of this chunk
            chunk_tags: Tags specific to this chunk
            page: Page number (for PDFs)
            section: Section name (for DOCX)
            
        Returns:
            LangChain Document object
        """
        document_tags = document_metadata.get("document_tags", []) or []
        merged_tags = sorted(set(document_tags) | set(chunk_tags))
        chunk_id = make_chunk_id(document_metadata, chunk_index, page=page, section=section)

        metadata = {
            "chunk_id": chunk_id,
            "chunk_index": chunk_index,
            "source": document_metadata["source"],
            "source_folder": document_metadata["source_folder"],
            "doc_type": document_metadata["doc_type"],
            "content_type": document_metadata.get("content_type", ""),
            "is_proposal": document_metadata.get("is_proposal", False),
            "page": page if page is not None else "",
            "section": section or "",
            "rfp_id": document_metadata.get("rfp_id", ""),
            "tags": merged_tags,
            "tags_text": ", ".join(merged_tags),
            "sensitive_data_obfuscated": document_metadata.get("sensitive_data_obfuscated", False),
            "language": "en",
            "file_hash": document_metadata.get("file_hash", ""),
            "pipeline_name": document_metadata.get("pipeline_name", ""),
            "processed_at": document_metadata.get("processed_at", ""),
            "script_version": SCRIPT_VERSION,
            "chunking_strategy": "langchain_recursive_character_splitter",
            "chunk_size_setting": config.CHUNK_SIZE,
            "chunk_overlap_setting": config.CHUNK_OVERLAP,
            "chunk_size_chars": len(content),
            "document_metadata": document_metadata,
        }

        return Document(page_content=content, metadata=safe_metadata(metadata))

DOC_BUILDER = KBDocumentBuilder()

# =============================================================================
# 8) PDF READER
# =============================================================================

def read_pdf(path: Path, source_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Extract cleaned text per PDF page while applying page skip rules."""
    results: List[Dict[str, Any]] = []
    rules = PDF_PAGE_OVERRIDES.get(source_name or path.name, {})
    skip_pages = set(rules.get("skip_pages", []))

    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            if page_num in skip_pages:
                continue
            text = clean(page.extract_text() or "")
            if len(text) >= config.MIN_CHUNK_LENGTH:
                results.append({"page": page_num, "text": text})
    return results

# =============================================================================
# 9) DOCX READER
# =============================================================================

def normalize_markdown_heading(text: str) -> str:
    """Clean a Markdown heading extracted from Docling output."""
    text = re.sub(r"^#{1,6}\s*", "", text).strip()
    text = re.sub(r"<!\-\-.*?\-\->", "", text)
    text = text.replace("**", "")
    text = unescape(text)
    text = re.sub(r"\\_", "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" |\t")

def extract_markdown_heading(line: str) -> Optional[Tuple[int, str]]:
    """Return a Markdown heading level and text when the line is a heading."""
    match = re.match(r"^(#{1,6})\s+(.*)$", line.strip())
    if not match:
        return None
    level = len(match.group(1))
    heading = normalize_markdown_heading(match.group(2))
    if not heading:
        return None
    return level, heading

def extract_table_section_heading(line: str) -> Optional[Tuple[int, str]]:
    """Detect section headings that Docling emitted as Markdown table rows."""
    stripped = line.strip()
    if not (stripped.startswith("|") and stripped.endswith("|")):
        return None
    if re.fullmatch(r"\|[\s:\-\|]+\|", stripped):
        return None

    cells = [normalize_markdown_heading(cell) for cell in stripped.strip("|").split("|")]
    cells = [cell for cell in cells if cell]
    if len(cells) < 2:
        return None

    first_cell = cells[0]
    second_cell = cells[1]

    if re.fullmatch(r"\d{1,2}", first_cell):
        title = re.sub(r"^#+\s*", "", second_cell).strip()
        if len(title) >= 4:
            return 1, f"{first_cell} {title}"

    joined = " ".join(cells)
    if "disclosure" in joined.lower() and "judgment" in joined.lower():
        return 1, joined

    return None

def section_should_skip(heading_path: List[str], skip_patterns: set) -> bool:
    """Check whether a heading path matches one of the configured skip rules."""
    path_text = " > ".join(heading_path).lower()
    return any(pattern in path_text for pattern in skip_patterns)

def read_docx_proposal(path: Path, source_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Convert the DOCX to Markdown, split it into kept sections, and return text."""
    if DocumentConverter is None:
        raise ImportError("Docling is required. Install it with: pip install docling")

    converter = DocumentConverter()
    file_name = source_name or path.name
    result = converter.convert(str(path))
    markdown_text = result.document.export_to_markdown()

    file_rules = DOCX_SECTION_OVERRIDES.get(file_name, {})
    extra_skip = {str(s).lower() for s in file_rules.get("skip_sections", [])}
    skip_patterns = DEFAULT_DOCX_SKIP_SECTION_PATTERNS | extra_skip

    stop_at_section = file_rules.get("stop_at_section")
    stop_at_section = str(stop_at_section).lower() if stop_at_section else None

    sections: List[Dict[str, Any]] = []
    heading_stack: List[str] = []
    current_heading_path: List[str] = []
    current_lines: List[str] = []
    current_keep = False
    section_index = 0

    def flush_current_section() -> None:
        """Save current section if it has content."""
        nonlocal section_index
        if not current_keep:
            return
        body = clean("\n".join(current_lines))
        if len(body) < config.MIN_CHUNK_LENGTH:
            return
        heading = " > ".join(current_heading_path).strip()
        if not heading:
            return
        sections.append({"heading": heading, "text": body, "section_index": section_index})
        section_index += 1

    for raw_line in markdown_text.splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()

        if not stripped:
            if current_keep:
                current_lines.append("")
            continue

        heading_info = extract_markdown_heading(stripped) or extract_table_section_heading(stripped)

        if heading_info:
            level, heading = heading_info
            flush_current_section()

            if stop_at_section and stop_at_section in heading.lower():
                current_heading_path = []
                current_lines = []
                current_keep = False
                break

            level = max(level, 1)
            if len(heading_stack) < level:
                heading_stack.extend([""] * (level - len(heading_stack)))
            heading_stack[level - 1] = heading
            heading_stack = heading_stack[:level]

            current_heading_path = [h for h in heading_stack if h]
            current_lines = []
            current_keep = not section_should_skip(current_heading_path, skip_patterns)
            continue

        if current_keep:
            current_lines.append(line)

    flush_current_section()

    print("\n===== DOCX SECTIONS =====")
    for s in sections:
        print(s["heading"])
    print("TOTAL SECTIONS:", len(sections))

    return sections

# =============================================================================
# 10) PII TRACKING (Development Only)
# =============================================================================

def track_pii_on_file(text: str, source: str, use_ner: bool = True):
    """Write a small length-only tracking log for debugging parsed text output."""
    try:
        logs_dir = ROOT_DIR / "logs"
        logs_dir.mkdir(exist_ok=True)

        safe_source = re.sub(r"[^a-zA-Z0-9_]", "_", source)[:40]
        filename = f"text_track_{safe_source}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        payload = {
            "source": source,
            "text_length_chars": len(text or ""),
            "tracked_at": datetime.now(timezone.utc).isoformat(),
            "note": "PII masking module is not used. This is a length-only development log.",
        }

        with open(logs_dir / filename, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        print(f"\n🔍 Text Tracking: {source}")
        print(f"   📄 Text length: {len(text or ''):,} characters")
        print(f"   📝 Saved tracking log: {filename}")

        return payload
    except Exception as e:
        print(f"   ⚠️  Text tracking failed: {e}")
        return None

# =============================================================================
# 11) PARALLEL PROCESSING PIPELINES
# =============================================================================

# -----------------------------------------------------------------------------
# 11.1) Pipeline A: Parallel PDF Processing
# -----------------------------------------------------------------------------

def process_single_pdf(args: Tuple) -> List[Document]:
    """Worker task: parse one PDF from ZIP bytes and return its Documents."""
    pdf_name, meta, entry_bytes, zip_hash = args
    
    try:
        # Create temporary file from ZIP bytes
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(entry_bytes)
            tmp_path = tmp.name
        
        try:
            # Build document metadata
            entry_hash = sha256_bytes(entry_bytes)
            doc_meta = build_document_metadata(
                source=pdf_name,
                source_folder="Beam_WeCloud_RAG_Documents.zip",
                doc_type=str(meta["doc_type"]),
                content_type="company_knowledge",
                pipeline_name=str(meta["pipeline_name"]),
                document_tags=list(meta["document_tags"]),
                file_hash_value=entry_hash,
                is_proposal=False,
                file_ext=".pdf",
            )
            doc_meta["container_zip_hash"] = zip_hash
            doc_meta["zip_entry_path"] = pdf_name
            doc_meta["sensitive_data_obfuscated"] = False
            
            # Extract and process PDF pages
            pages = read_pdf(Path(tmp_path), source_name=pdf_name)
            
            # Build Document objects for each page
            documents = []
            for page_item in pages:
                page_num = page_item["page"]
                page_text = page_item["text"]
                
                # Chunk the page text
                for local_chunk_index, piece in enumerate(CHUNKER.split(page_text)):
                    chunk_index = (page_num * 1000) + local_chunk_index
                    chunk_tags = auto_tag(piece, list(meta["document_tags"]))
                    
                    doc = DOC_BUILDER.build(
                        content=piece,
                        document_metadata=doc_meta,
                        chunk_index=chunk_index,
                        chunk_tags=chunk_tags,
                        page=page_num,
                        section="",
                    )
                    documents.append(doc)
            
            return documents
            
        finally:
            # Clean up temporary file
            os.unlink(tmp_path)
            
    except Exception as e:
        print(f"❌ Failed to process {pdf_name}: {e}")
        return []

def pipeline_a_parallel(oai_client: Optional[OpenAI] = None) -> List[Document]:
    """Read target company PDFs from the ZIP and process them in parallel."""
    print("\n📄 Pipeline A: Processing PDFs in parallel")
    
    documents: List[Document] = []
    
    # Locate the ZIP file
    company_zip = resolve_zip_path(COMPANY_ZIP, "Beam_WeCloud_RAG_Documents")
    if company_zip is None:
        print(f"  ❌ ZIP not found in: {SOURCE_DATA_DIR}")
        return documents
    
    print(f"  📦 ZIP: {company_zip.name}")
    zip_hash = file_hash(company_zip)
    
    # Prepare tasks for parallel processing
    tasks = []
    with zipfile.ZipFile(company_zip) as zf:
        for pdf_name, meta in COMPANY_PDF_FILES.items():
            entry_path = find_zip_entry(zf, pdf_name)
            if entry_path is None:
                print(f"  ⚠️  Not found in ZIP: {pdf_name}")
                continue
            
            entry_bytes = zf.read(entry_path)
            tasks.append((pdf_name, meta, entry_bytes, zip_hash))
    
    if not tasks:
        print("  ⚠️  No PDF files to process")
        return documents
    
    print(f"  📄 Processing {len(tasks)} PDF files in parallel on {MAX_WORKERS_CPU} cores")
    
    # Submit one PDF-processing task per target PDF
    with ProcessPoolExecutor(max_workers=MAX_WORKERS_CPU) as executor:
        # Submit all tasks
        futures = {executor.submit(process_single_pdf, task): task[0] for task in tasks}
        
        # Collect results with progress bar
        with tqdm(total=len(futures), desc="  Processing PDFs") as pbar:
            for future in as_completed(futures):
                pdf_name = futures[future]
                try:
                    result = future.result(timeout=120)  # per-file timeout
                    documents.extend(result)
                    pbar.set_postfix({'✅': pdf_name[:30]})
                except Exception as e:
                    print(f"\n  ❌ Failed {pdf_name}: {e}")
                pbar.update(1)
    
    print(f"  ✅ Processed {len(documents)} documents from PDFs")
    return documents

# -----------------------------------------------------------------------------
# 11.2) Pipeline B: Parallel DOCX Processing
# -----------------------------------------------------------------------------

def process_single_docx(args: Tuple) -> List[Document]:
    """
    Worker task: parse the USask DOCX proposal and return its Documents.

    The DOCX file is only one file, so file-level parallelism does not help much.
    The expensive part is LLM masking/tagging per section, so this function
    parallelizes section processing with ThreadPoolExecutor.
    """
    entry, entry_bytes, zip_hash, oai_client = args

    try:
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(entry_bytes)
            tmp_path = tmp.name

        try:
            fname = Path(entry.filename).name
            folder_name = Path(entry.filename).parent.name or "Tenders"

            is_usask = (
                "usask" in folder_name.lower()
                or "730126" in folder_name.lower()
                or "730126" in fname.lower()
            )
            rfp_id = USASK_RFP_ID if is_usask else re.sub(r"[^\w]", "_", folder_name)[:50]

            entry_hash = sha256_bytes(entry_bytes)

            doc_meta = build_document_metadata(
                source=fname,
                source_folder="Tenders - Raw and Annotated.zip",
                doc_type="proposal",
                content_type="proposal_response",
                pipeline_name="tender_usask_docx_proposal_pipeline",
                document_tags=["proposal", "usask", "genai", "education", "government"],
                file_hash_value=entry_hash,
                rfp_id=rfp_id,
                is_proposal=True,
                file_ext=".docx",
            )

            doc_meta["container_zip_hash"] = zip_hash
            doc_meta["zip_entry_path"] = entry.filename
            doc_meta["sensitive_data_obfuscated"] = bool(oai_client)

            sections = read_docx_proposal(Path(tmp_path), source_name=fname)
            full_text = "\n\n".join(s["text"] for s in sections)
            track_pii_on_file(full_text, fname, use_ner=True)

            def process_section(
                section_payload: Tuple[int, Dict[str, Any]]
            ) -> Tuple[int, List[Document]]:
                """
                Process one DOCX section.

                Steps:
                - Optionally mask sensitive data with the LLM.
                - Apply regex validation after masking.
                - Generate section-level tags.
                - Chunk the masked section text.
                - Convert chunks into LangChain Documents.
                """
                section_index, section = section_payload
                section_text = section["text"]

                if oai_client:
                    print(f"    [mask] Running LLM masking on section: {section['heading']}")
                    llm_masked = _mask_via_llm(section_text, oai_client)
                    masked_section_text = _regex_validate_and_fix(llm_masked)
                    section_tags = llm_tag(
                        masked_section_text,
                        oai_client,
                        base_tags=doc_meta["document_tags"],
                    )
                else:
                    masked_section_text = section_text
                    section_tags = auto_tag(
                        masked_section_text,
                        doc_meta["document_tags"],
                    )

                section_documents: List[Document] = []

                for local_chunk_index, piece in enumerate(CHUNKER.split(masked_section_text)):
                    chunk_index = (
                        section.get("section_index", section_index) * 1000
                    ) + local_chunk_index

                    chunk_tags = auto_tag(piece, section_tags)

                    section_documents.append(
                        DOC_BUILDER.build(
                            content=piece,
                            document_metadata=doc_meta,
                            chunk_index=chunk_index,
                            chunk_tags=chunk_tags,
                            page=None,
                            section=section["heading"],
                        )
                    )

                print(
                    f"    [section] {section['heading'][:70]} "
                    f"-> {len(section_documents)} chunk(s)"
                )

                return section_index, section_documents

            documents: List[Document] = []
            section_results: List[Tuple[int, List[Document]]] = []

            if sections:
                max_section_workers = min(DOCX_SECTION_WORKERS, len(sections))

                print(
                    f"    [parallel] Processing {len(sections)} DOCX sections "
                    f"with {max_section_workers} threads"
                )

                with ThreadPoolExecutor(max_workers=max_section_workers) as executor:
                    futures = [
                        executor.submit(process_section, (section_index, section))
                        for section_index, section in enumerate(sections)
                    ]

                    for future in as_completed(futures):
                        section_results.append(future.result())

                # Restore deterministic section order after parallel execution.
                for _, section_docs in sorted(section_results, key=lambda item: item[0]):
                    documents.extend(section_docs)

            print(f"  ✅ DOCX chunks generated: {len(documents)}")
            return documents

        finally:
            os.unlink(tmp_path)

    except Exception as e:
        print(f"❌ Failed to process DOCX {entry.filename}: {e}")
        return []

def pipeline_b_parallel(oai_client: Optional[OpenAI] = None) -> List[Document]:
    """Read the target DOCX proposal from the ZIP and convert it into Documents."""
    print("\n📝 Pipeline B: Processing DOCX in parallel")
    
    documents: List[Document] = []
    
    # Locate the ZIP file
    tenders_zip = resolve_zip_path(TENDERS_ZIP, "Tenders - Raw and Annotated")
    if tenders_zip is None:
        print(f"  ❌ ZIP not found in: {SOURCE_DATA_DIR}")
        return documents
    
    print(f"  📦 ZIP: {tenders_zip.name}")
    zip_hash = file_hash(tenders_zip)
    
    # Find DOCX files in ZIP
    docx_tasks = []
    with zipfile.ZipFile(tenders_zip) as zf:
        for entry in zf.infolist():
            if entry.is_dir():
                continue
            fname = Path(entry.filename).name
            ext = Path(fname).suffix.lower()
            if ext == ".docx" and "copy of cp-730126" in fname.lower():
                entry_bytes = zf.read(entry.filename)
                docx_tasks.append((entry, entry_bytes, zip_hash, oai_client))
    
    if not docx_tasks:
        print("  ⚠️  No DOCX files to process")
        return documents
    
    print(f"  📄 Found {len(docx_tasks)} DOCX file(s)")
    
    # Process each DOCX file
    for task in docx_tasks:
        result = process_single_docx(task)
        documents.extend(result)
        print(f"  ✅ Processed: {Path(task[0].filename).name}")
    
    return documents

# -----------------------------------------------------------------------------
# 11.3) Pipeline C: Parallel Excel Scope Processing
# -----------------------------------------------------------------------------

SCOPE_HEADER_KEYWORDS = [
    "scope of services",
    "scope of service",
    "scope of work",
    "project scope",
    "requirements",
]

SECTION_STOP_KEYWORDS = [
    "deliverables",
    "timeline",
    "timelines",
    "pricing",
    "submission",
    "evaluation",
    "appendix",
    "references",
]

def normalize_excel_text(text: Any) -> str:
    """Normalize Excel cell text for matching headers and stop words."""
    return re.sub(r"\s+", " ", str(text or "").strip().lower())

def clean_scope_cell_text(text: Any) -> str:
    """Clean a candidate Excel scope cell before capability extraction."""
    if text is None:
        return ""
    cleaned = str(text).strip()
    if "|" in cleaned:
        cleaned = cleaned.split("|")[0].strip()
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()

def find_scope_header(ws) -> Tuple[Optional[int], Optional[int], Optional[str]]:
    """Find the worksheet cell that marks the start of a scope section."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 6) + 1):
            value = ws.cell(row=row, column=col).value
            text = normalize_excel_text(value)
            if any(k in text for k in SCOPE_HEADER_KEYWORDS):
                return row, col, str(value)
    return None, None, None

def extract_scope_openpyxl(path: Path) -> Tuple[str, Optional[Dict[str, Any]]]:
    """Extract the scope block and extraction metadata from one Excel workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)

    for ws in wb.worksheets:
        header_row, header_col, header_value = find_scope_header(ws)

        if header_row is None:
            continue

        candidate_cells = []

        for r in range(header_row + 1, min(ws.max_row, header_row + 15) + 1):
            candidate_cells.append((r, header_col))
            if header_col != 1:
                candidate_cells.append((r, 1))

        for r, c in candidate_cells:
            raw_value = ws.cell(row=r, column=c).value
            cleaned = clean_scope_cell_text(raw_value)
            norm = normalize_excel_text(cleaned)

            if not cleaned:
                continue

            if len(norm) < 100 and any(k in norm for k in SECTION_STOP_KEYWORDS):
                continue

            if len(cleaned) > 120 or re.search(r"\n\s*1\.|1\.", cleaned):
                return cleaned, {
                    "sheet": ws.title,
                    "scope_header_row": header_row,
                    "scope_header_col": header_col,
                    "scope_header_value": header_value,
                    "scope_cell_row": r,
                    "scope_cell_col": c,
                    "scope_cell_address": ws.cell(row=r, column=c).coordinate,
                }

    return "", None

def build_scope_capability_prompt(scope_text: str) -> str:
    """Create the LLM prompt that turns tender scope text into capability JSON."""
    return f"""
You are preparing clean structured data for a Beam Data RAG knowledge base.

You will receive ONLY the Scope text extracted from an Excel tender/project file.

Your task:
1. Extract reusable capabilities from the Scope only.
2. Infer the industry only if it is clearly supported by the Scope text.
3. Remove or obfuscate sensitive information.
4. Produce a detailed capability summary that preserves all relevant technical capabilities found in the Scope.

Important rules:
Only extract capabilities that represent reusable technical services,
platforms, architectures, integrations, infrastructure, AI methods,
or consulting offerings.

Many scope items may be written as project phases or implementation activities.
When this occurs, identify and extract the underlying reusable technical capabilities,
platforms, architectures, AI methods, data capabilities, and infrastructure implied by those activities.
Focus on WHAT can be delivered, not HOW the project is organized.

Do not extract:
- project phases
- business outcomes
- planning activities
- management activities
- operational tasks
- generic benefits
- reporting outputs

Do not create capabilities that are not explicitly supported by the scope.

Return valid JSON only with this exact structure:

{{
  "capability_text": "",
  "metadata": {{
    "source_type": "tender_scope_capabilities",
    "evidence_type": "inferred_from_tender_scope",
    "confidence": "medium",
    "industry": "",
    "capabilities": []
  }},
  "capability_details": [
    {{
      "capability": "",
      "description": "",
      "generic_evidence": ""
    }}
  ]
}}

The capability_text should provide a comprehensive capability summary rather than a brief paragraph.
Multiple paragraphs are allowed if needed to preserve all important technical details.

The summary should begin with:
"Beam Data Capability Areas in <Industry>:"

Then describe all relevant capabilities supported by the scope in a clear, reusable format suitable for a RAG knowledge base.

Do not use the word "chunk" in the output text.
Do not summarize too aggressively.
Do not omit specific technical terms if they are explicitly mentioned in the scope.

Scope text:
{scope_text}
"""

def parse_llm_json(text: str) -> Dict[str, Any]:
    """Parse JSON from an LLM response, including markdown-wrapped JSON."""
    text = text.strip()
    text = re.sub(r"^```json\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^```\s*", "", text)
    text = re.sub(r"\s*```$", "", text)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if not match:
            raise
        return json.loads(match.group(0))

def extract_capabilities_with_llm(scope_text: str, client: OpenAI) -> Dict[str, Any]:
    """Call the LLM and return parsed capability extraction JSON."""
    prompt = build_scope_capability_prompt(scope_text)
    resp = client.responses.create(
        model=config.LLM_MODEL,
        input=prompt,
    )
    return parse_llm_json(resp.output_text)

def process_single_excel(args: Tuple) -> Optional[Dict[str, Any]]:
    """Worker task: extract one Excel scope and call the LLM for capabilities."""
    entry, entry_bytes, zip_hash, oai_client, tender_files = args
    
    fname = Path(entry.filename).name
    file_stem = normalize_file_stem(fname)
    
    # Check if this file is targeted
    if file_stem not in tender_files:
        return None
    
    print(f"\n  [excel] Processing: {entry.filename}")
    
    try:
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(entry_bytes)
            tmp_path = tmp.name
        
        try:
            # Extract scope text
            scope_text, extraction_info = extract_scope_openpyxl(Path(tmp_path))
            
            if not scope_text.strip():
                print(f"  [skip] No scope found: {fname}")
                return None
            
            # Extract capabilities using LLM
            result = extract_capabilities_with_llm(scope_text, oai_client)
            
            capability_text = result.get("capability_text", "").strip()
            result_meta = result.get("metadata", {}) or {}
            capabilities = result_meta.get("capabilities", []) or []
            industry = result_meta.get("industry", "") or ""
            
            if not capability_text:
                print(f"  [skip] Empty capability_text: {fname}")
                return None
            
            # Track progress
            track_pii_on_file(capability_text, fname, use_ner=True)
            
            return {
                "fname": fname,
                "entry": entry,
                "entry_bytes": entry_bytes,
                "zip_hash": zip_hash,
                "scope_text": scope_text,
                "capability_text": capability_text,
                "capabilities": capabilities,
                "industry": industry,
                "result_meta": result_meta,
                "extraction_info": extraction_info,
                "entry_hash": sha256_bytes(entry_bytes),
            }
            
        finally:
            os.unlink(tmp_path)
            
    except Exception as e:
        print(f"  [error] Excel pipeline failed for {fname}: {e}")
        return None

def pipeline_c_excel_scopes(oai_client: Optional[OpenAI] = None) -> List[Document]:
    """Process target Excel scopes in parallel and convert capability outputs to Documents."""
    documents: List[Document] = []
    
    target_count = len(TENDER_EXCEL_TARGET_FILES)
    matched_files: List[str] = []
    processed_files: List[str] = []
    skipped_files: List[str] = []
    error_files: List[str] = []

    if oai_client is None:
        print("  [skip] Excel capability pipeline requires OpenAI client")
        return documents

    print("\n📊 Pipeline C: Processing Excel files in parallel")
    
    # Locate the ZIP file
    tenders_zip = resolve_zip_path(TENDERS_ZIP, "Tenders - Raw and Annotated")
    if tenders_zip is None:
        print(f"  [skip] Tenders ZIP not found in: {SOURCE_DATA_DIR}")
        return documents

    print(f"  📦 ZIP: {tenders_zip.name}")
    zip_hash = file_hash(tenders_zip)
    
    # Collect Excel files for parallel processing
    excel_tasks = []
    with zipfile.ZipFile(tenders_zip) as zf:
        for entry in zf.infolist():
            if entry.is_dir():
                continue
            fname = Path(entry.filename).name.strip()
            ext = Path(fname).suffix.lower()
            
            if ext != ".xlsx":
                continue
            
            file_stem = normalize_file_stem(fname)
            if file_stem in TENDER_EXCEL_TARGET_FILES:
                entry_bytes = zf.read(entry.filename)
                excel_tasks.append((entry, entry_bytes, zip_hash, oai_client, TENDER_EXCEL_TARGET_FILES))
                matched_files.append(fname)
    
    if not excel_tasks:
        print("  ⚠️  No Excel files to process")
        return documents
    
    print(f"  📊 Processing {len(excel_tasks)} Excel files in parallel with {MAX_WORKERS_IO} threads")
    
    # Submit one Excel-processing task per target workbook
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_IO) as executor:
        # Submit all tasks
        futures = {executor.submit(process_single_excel, task): task for task in excel_tasks}
        
        # Collect results with progress bar
        with tqdm(total=len(futures), desc="  Processing Excel files") as pbar:
            for future in as_completed(futures):
                task = futures[future]
                fname = Path(task[0].filename).name
                
                try:
                    result = future.result(timeout=120)
                    
                    if result:
                        # Build documents from the result
                        doc_meta = build_document_metadata(
                            source=result["fname"],
                            source_folder="Tenders - Raw and Annotated.zip",
                            doc_type="tender_scope_capabilities",
                            content_type="capability_summary",
                            pipeline_name="tender_excel_scope_capabilities_pipeline",
                            document_tags=["tender", "tender-scope", "capabilities", "beamdata"],
                            file_hash_value=result["entry_hash"],
                            rfp_id=re.sub(r"[^\w]+", "_", Path(result["entry"].filename).parent.name)[:80],
                            is_proposal=False,
                            file_ext=".xlsx",
                        )
                        
                        # Add metadata
                        doc_meta["container_zip_hash"] = result["zip_hash"]
                        doc_meta["zip_entry_path"] = result["entry"].filename
                        doc_meta["sensitive_data_obfuscated"] = False
                        doc_meta["industry"] = result["industry"]
                        doc_meta["capabilities"] = result["capabilities"]
                        doc_meta["capabilities_text"] = ", ".join(map(str, result["capabilities"]))
                        doc_meta["source_type"] = result["result_meta"].get("source_type", "tender_scope_capabilities")
                        doc_meta["evidence_type"] = result["result_meta"].get("evidence_type", "inferred_from_tender_scope")
                        doc_meta["confidence"] = result["result_meta"].get("confidence", "medium")
                        doc_meta["scope_length_chars"] = len(result["scope_text"])
                        doc_meta["capability_text_length_chars"] = len(result["capability_text"])
                        
                        if result["extraction_info"]:
                            doc_meta["scope_sheet"] = result["extraction_info"].get("sheet", "")
                            doc_meta["scope_cell"] = result["extraction_info"].get("scope_cell_address", "")
                        
                        # Build tags
                        base_tags = ["tender", "tender-scope", "capabilities", "beamdata"]
                        if result["industry"]:
                            industry_tag = slugify_tag(result["industry"])
                            if industry_tag:
                                base_tags.append(industry_tag)
                        
                        for cap in result["capabilities"]:
                            cap_tag = slugify_tag(cap)
                            if cap_tag and len(cap_tag) <= 50:
                                base_tags.append(cap_tag)
                        
                        base_tags = sorted(set(base_tags))
                        chunk_tags = auto_tag(result["capability_text"], base_tags)
                        
                        # Chunk and build documents
                        for local_chunk_index, piece in enumerate(CHUNKER.split(result["capability_text"])):
                            documents.append(DOC_BUILDER.build(
                                content=piece,
                                document_metadata=doc_meta,
                                chunk_index=local_chunk_index,
                                chunk_tags=auto_tag(piece, chunk_tags),
                                page=None,
                                section="Tender Scope Capabilities",
                            ))
                        
                        processed_files.append(result["fname"])
                        pbar.set_postfix({'✅': result["fname"][:20]})
                    
                except Exception as e:
                    print(f"\n  [error] Excel pipeline failed for {fname}: {e}")
                    error_files.append(f"{fname} — {e}")
                
                pbar.update(1)
    
    # Summary report
    unmatched_targets = sorted(TENDER_EXCEL_TARGET_FILES - {normalize_file_stem(f) for f in matched_files})
    
    print("\n" + "=" * 80)
    print("EXCEL PIPELINE SUMMARY")
    print("=" * 80)
    print(f"Target Excel files       : {target_count}")
    print(f"Matched target files     : {len(matched_files)}")
    print(f"Processed successfully   : {len(processed_files)}")
    print(f"Skipped after matching   : {len(skipped_files)}")
    print(f"Errors                   : {len(error_files)}")
    print(f"Documents generated      : {len(documents)}")
    
    if processed_files:
        print("\nProcessed files:")
        for name in processed_files:
            print(f"  - {name}")
    
    if skipped_files:
        print("\nSkipped files:")
        for name in skipped_files:
            print(f"  - {name}")
    
    if error_files:
        print("\nFiles with errors:")
        for name in error_files:
            print(f"  - {name}")
    
    if unmatched_targets:
        print("\nTarget files not found in ZIP:")
        for name in unmatched_targets:
            print(f"  - {name}")
    
    print("=" * 80)
    
    return documents

# =============================================================================
# 12) SAVE / INDEX
# =============================================================================

def dedupe_documents(documents: List[Document]) -> List[Document]:
    """Remove exact duplicate chunks after normalizing whitespace and case."""
    seen = set()
    deduped: List[Document] = []
    for doc in documents:
        norm = re.sub(r"\s+", " ", doc.page_content.lower().strip())
        if norm not in seen:
            seen.add(norm)
            deduped.append(doc)
    removed = len(documents) - len(deduped)
    if removed:
        print(f"\n  Deduplication: removed {removed} duplicate chunks")
    return deduped

def save_documents_json(documents: List[Document]) -> None:
    """Write page_content and metadata for all Documents to kb_documents.json."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = [{"page_content": d.page_content, "metadata": d.metadata} for d in documents]
    with open(DOCS_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"Saved documents JSON -> {DOCS_JSON}")

def print_stats(documents: List[Document]) -> None:
    """Print document counts by type and the most common tags."""
    print(f"\nTotal LangChain Documents: {len(documents)}")

    by_type: Dict[str, int] = {}
    for doc in documents:
        doc_type = str(doc.metadata.get("doc_type", "unknown"))
        by_type[doc_type] = by_type.get(doc_type, 0) + 1

    print("\nDocuments by doc_type:")
    for doc_type, count in sorted(by_type.items()):
        print(f"  {doc_type:<25} {count}")

    all_tags: Dict[str, int] = {}
    for doc in documents:
        tags_text = str(doc.metadata.get("tags_text", ""))
        for tag in [t.strip() for t in tags_text.split(",") if t.strip()]:
            all_tags[tag] = all_tags.get(tag, 0) + 1

    top_tags = sorted(all_tags.items(), key=lambda x: -x[1])[:10]
    print("\nTop tags:")
    print("  " + ", ".join(f"{tag}({count})" for tag, count in top_tags))




def print_file_chunk_report(documents: List[Document]) -> None:
    """
    Print chunk counts per source file.

    This is used to compare sequential vs parallel builds and confirm that
    both versions generate the same number of chunks per file and doc_type.
    """

    print("\nChunks by source file:")
    print("-" * 100)
    print(f"{'Source':<65} {'Doc Type':<28} {'Chunks':>6}")
    print("-" * 100)

    counts: Dict[Tuple[str, str], int] = {}

    for doc in documents:
        source = str(doc.metadata.get("source", "unknown"))
        doc_type = str(doc.metadata.get("doc_type", "unknown"))
        key = (source, doc_type)
        counts[key] = counts.get(key, 0) + 1

    for (source, doc_type), count in sorted(counts.items(), key=lambda item: (item[0][1], item[0][0])):
        print(f"{source[:65]:<65} {doc_type[:28]:<28} {count:>6}")

    print("-" * 100)


def print_docx_section_chunk_report(documents: List[Document]) -> None:
    """
    Print chunk counts per DOCX section.

    Helpful for debugging whether parallel section processing changed the
    number of generated chunks for any section.
    """

    section_counts: Dict[str, int] = {}

    for doc in documents:
        if doc.metadata.get("doc_type") != "proposal":
            continue

        section = str(doc.metadata.get("section", ""))
        section_counts[section] = section_counts.get(section, 0) + 1

    if not section_counts:
        return

    print("\nDOCX proposal chunks by section:")
    print("-" * 100)
    print(f"{'Section':<90} {'Chunks':>6}")
    print("-" * 100)

    for section, count in sorted(section_counts.items()):
        print(f"{section[:90]:<90} {count:>6}")

    print("-" * 100)


def format_seconds(seconds: float) -> str:
    """Format elapsed seconds for build logs."""
    if seconds < 60:
        return f"{seconds:.2f} sec"
    minutes, sec = divmod(seconds, 60)
    return f"{int(minutes)} min {sec:.2f} sec"

# =============================================================================
# 13) MAIN
# =============================================================================

def main() -> None:
    """Run the full KB build, print timing metrics, and optionally build Chroma."""
    total_start = time.perf_counter()
    timing: Dict[str, float] = {}

    parser = argparse.ArgumentParser(
        description="Build LangChain Knowledge Base with parallel processing"
    )

    parser.add_argument(
        "--reset",
        action="store_true",
        help="Delete and rebuild the LangChain Chroma DB",
    )

    parser.add_argument(
        "--no-index",
        action="store_true",
        help="Only write kb_documents.json, do not build Chroma",
    )

    args = parser.parse_args()

    print("=" * 70)
    print(f"  Building LangChain Knowledge Base — {SCRIPT_VERSION}")
    print("=" * 70)
    print(f"  ⚡ Parallel Processing Enabled (CPU: {CPU_CORES} cores)")

    # Initialize OpenAI client
    oai_client = None
    if config.OPENAI_API_KEY:
        oai_client = OpenAI(api_key=config.OPENAI_API_KEY)
        print("  🤖 LLM tagging: ENABLED")
        print("  🔒 LLM masking: ENABLED")
    else:
        print("  🤖 LLM tagging: DISABLED — keyword tags only")
        print("  🔒 LLM masking: DISABLED")

    print("\n" + "=" * 70)
    print("  STARTING PARALLEL PROCESSING PIPELINES")
    print("=" * 70)

    # Pipeline A: company PDFs
    print("\n[1/3] Pipeline A — Company PDFs (Parallel)")
    step_start = time.perf_counter()
    docs_a = pipeline_a_parallel(oai_client)
    timing["Pipeline A - PDFs"] = time.perf_counter() - step_start
    print(f"      ✅ {len(docs_a)} documents generated")
    print(f"      ⏱ Time: {format_seconds(timing['Pipeline A - PDFs'])}")

    # Pipeline B: USask DOCX proposal
    print("\n[2/3] Pipeline B — USask DOCX Proposal")
    step_start = time.perf_counter()
    docs_b = pipeline_b_parallel(oai_client)
    timing["Pipeline B - DOCX"] = time.perf_counter() - step_start
    print(f"      ✅ {len(docs_b)} documents generated")
    print(f"      ⏱ Time: {format_seconds(timing['Pipeline B - DOCX'])}")

    # Pipeline C: Excel tender scope capability summaries
    print("\n[3/3] Pipeline C — Excel Tender Scope Capabilities (Parallel)")
    step_start = time.perf_counter()
    docs_c = pipeline_c_excel_scopes(oai_client)
    timing["Pipeline C - Excel"] = time.perf_counter() - step_start
    print(f"      ✅ {len(docs_c)} documents generated")
    print(f"      ⏱ Time: {format_seconds(timing['Pipeline C - Excel'])}")

    # Combine outputs and remove duplicate chunks
    all_docs = docs_a + docs_b + docs_c
    print(f"\n📊 Total documents before deduplication: {len(all_docs)}")
    step_start = time.perf_counter()
    documents = dedupe_documents(all_docs)
    timing["Deduplication"] = time.perf_counter() - step_start

    # Statistics
    proposals = sum(1 for d in docs_b if d.metadata.get("is_proposal") is True)
    print(f"      📝 {proposals} proposal documents")

    print_stats(documents)
    print_file_chunk_report(documents)
    print_docx_section_chunk_report(documents)

    # Save the intermediate JSON used by evaluation and debugging
    step_start = time.perf_counter()
    save_documents_json(documents)
    timing["Save JSON"] = time.perf_counter() - step_start

    # Build or refresh the Chroma vector index unless --no-index is provided
    if not args.no_index:
        print("\n🔍 Building Chroma index with LangChain...")
        step_start = time.perf_counter()

        store = ChromaVectorStore(
            persist_dir=config.CHROMA_DIR,
            collection_name=config.COLLECTION_NAME,
            embedding_model=config.EMBEDDING_MODEL,
        )

        vectorstore = store.build(
            documents=documents,
            reset=args.reset,
        )

        timing["Chroma Indexing"] = time.perf_counter() - step_start
        print(f"✅ Chroma ready -> {config.CHROMA_DIR}")
        print(f"   Collection: {config.COLLECTION_NAME}")
        print(f"   Vector count: {vectorstore._collection.count()}")
        print(f"   ⏱ Time: {format_seconds(timing['Chroma Indexing'])}")

    print("\n" + "=" * 70)
    print("  ✅ BUILD COMPLETE")
    print("=" * 70)
    print(
        "🚀 Next: Run the API with: "
        "uvicorn src.api_langchain:app --reload"
    )
    total_time = time.perf_counter() - total_start

    print("\nPerformance Summary:")
    print(f"  - Processed {len(documents)} documents")
    print(f"  - Total build time: {format_seconds(total_time)}")
    for step_name, elapsed in timing.items():
        print(f"  - {step_name}: {format_seconds(elapsed)}")
    print(f"  - Used {CPU_CORES} CPU cores for PDF process workers")
    print(f"  - Used {MAX_WORKERS_IO} threads for Excel/LLM workers")
    print(f"  - Used {DOCX_SECTION_WORKERS} threads for DOCX section workers")

if __name__ == "__main__":
    main()